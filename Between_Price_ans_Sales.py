# сравниваем 2 файла
# в первом файле отчёт с информацией о цене на товар и авторе этой цены
# во втором файле отчёт по продажам товара на определённую сумму
# нужно создать файл с таблицей продаж указать авторов цены каждой позиции

import openpyxl
import os

path = os.chdir("D:\\Dropbox\\LeessonS\\WB_Photo\\Between") # устанавливаем рабочий каталог, где лежат файлы для сравнения
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# заголовки в новом файле будут такими
table_tytles = ('Дата продажи', 'Товар', 'кол-во','Цена', 'Автор цены')
for col_num in range (len(table_tytles)):
    new_ws.cell(row = 1, column=col_num+1).value = table_tytles[col_num]
#  переменная 'i' - используется в адресации строк нового файла
i = 2
wb_sale = openpyxl.open('D:\\Dropbox\\LeessonS\\WB_Photo\\Between\\Sale.xlsx')  # открываем книгу, ссылаясь на текущий файл
wb_sale.active = 0 # устанавливаем активный лист
ws_sale = wb_sale.active

wb_price = openpyxl.open('D:\\Dropbox\\LeessonS\\WB_Photo\\Between\\Price.xlsx')  # открываем книгу, ссылаясь на текущий файл
wb_price.active = 0  # устанавливаем активный лист (определён заранее)
ws_price = wb_price.active

# определяем списки хранения данных из таблицы с ценами / для авторов / для названия товара
price_autor = []
price_product = []
# создаём список для хранения авторов цены
users = []

# заполняем списки считывая файл Price
for row_in_price in range(2, ws_price.max_row + 1):
    product = ws_price[row_in_price][2].value
    autor = ws_price[row_in_price][1].value
    price_autor.append(autor)
    price_product.append(product)
# определяем длину каждого списка (они равны по длине)
large_list = len(price_autor)

# считываем файл с продажами и формируем новый файл с аналогичными колонками
for row in range(2, ws_sale.max_row + 1):
    print(f"Working in {row} of {ws_sale.max_row+1}")
    new_ws.cell(row=i, column=1).value = sale_date = ws_sale[row][0].value
    new_ws.cell(row=i, column=2).value = sale_product = ws_sale[row][1].value
    new_ws.cell(row=i, column=3).value = sale_count = ws_sale[row][2].value
    new_ws.cell(row=i, column=4).value = sale_price = ws_sale[row][3].value
    # получив данные о товаре проводим перебор списка товаров

    for pro in range(large_list):
        # и в случае равенства названий товаров


        if sale_product == price_product[pro]:
            # добавляем авторов цены в список users
            users.append(price_autor[pro])

    # записываем в отдельную ячейку множество (set) авторов, "приложивших руку" к цене
    new_ws.cell(row=i, column=5).value =str(set(users))
    # обнуляем список для использования на следующей итерации проверки
    users = []

    # увеличиваем значение строки для записи на следующей строке листа
    i += 1
    # далее возврат к началу цикла или окончание цикла и запись итоговой информации в файл
# запись в файл
print("Saved New_File and exit")
new_wb.save("between.xlsx")
# закрываем книгу для исключения ошибки совместного доступа
new_wb.close()
# exit()