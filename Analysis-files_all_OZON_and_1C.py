# Программа сранивает два файла (весь каталог ОЗОН и выгрузку из 1с)
# и добавляет в отдельную колонку ШК_озона, если он есть на ОЗОНе


import openpyxl
import os

# устанавливаем рабочий каталог, где лежат файлы с OZON
# path = os.chdir("C:\\Users\\AlexMiki\\Desktop\\ПРОЕКТЫ\\Переезд OZON to WB\\OZON товары")
# path = os.chdir("F:\\Dropbox\\LeessonS\\WB_Photo\\OZONE")
path = os.chdir("D:\\Dropbox\\LeessonS\\1С\\")

# РАБОТА С СОЗДАВАЕМЫМ ФАЙЛОМ
# инициализация НОВОГО Excel-файла и листа для записи
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# перечисляем фиксированные имена полей будущей таблицы
table_tytles = ('ID_1C','ID_site_OSCOMP','BAR_OZONE_in_1C','Наименование',
                'Полное Наименование','BAR_OZON_in_OZONE')
# записываем поля в файл
for col_num in range (0, 5):
    new_ws.cell(row = 1, column=col_num+1).value = table_tytles[col_num]
#  переменная 'i' - используется в адресации строк (начнём со 2-й строки)
i = 2
# определение рабочих файлов
    # с данными 1С:
wb_1C = openpyxl.open('KATALOG_1C.xlsx')
wb_1C.active = 0 # устанавливаем активный лист
ws_1C = wb_1C.active
    # с данными OZON:
wb_OZONE = openpyxl.open('ALL_catalog_from_OZON_in_one_file.xlsx')
wb_OZONE.active = 0 # устанавливаем активный лист
ws_OZONE = wb_OZONE.active
# считываем список значений ячеек с листа, начиная с 4-й строки
# заголовки пропускаются
ozone_cat = {}
for row_OZONE in range (2, ws_OZONE.max_row+1):
    id_site_in_OZONE = ws_OZONE[row_OZONE][0].value
    bar_ozone_in_ozone = ws_OZONE[row_OZONE][4].value
    ozone_cat [str(id_site_in_OZONE)] = str (bar_ozone_in_ozone)
    # print(row_OZONE, id_site_in_OZONE, bar_ozone_in_ozone)

for row_1C in range (1, ws_1C.max_row+1):
    id_1C = ws_1C[row_1C][0].value
    id_site_OSCOMP = ws_1C[row_1C][1].value
    BAR_OZONE_in_1C = ws_1C[row_1C][2].value
    name = ws_1C[row_1C][3].value
    full_name = ws_1C[row_1C][4].value
    # полученные значения  записываются в столбцы текущей строки (i)
    # у новой рабочей книги 'ws'
    new_ws.cell(row=i, column=1).value = id_1C
    new_ws.cell(row=i, column=2).value = id_site_OSCOMP
    new_ws.cell(row=i, column=3).value = BAR_OZONE_in_1C
    new_ws.cell(row=i, column=4).value = name
    new_ws.cell(row=i, column=5).value = full_name
    if id_site_OSCOMP in ozone_cat.keys():
        new_ws.cell(row=i, column=6).value = ozone_cat[id_site_OSCOMP]
    # увеличиваем значение строки для записи на следующей строке листа
    print(row_1C, ' of \t', ws_1C.max_row+1, id_site_OSCOMP)
    i += 1
# по окончании обработки всех файлов сохраняем изменения в созданном файле
new_wb.save("Analisys_1C_and_OZON.xlsx")
# закрываем книгу для исключения ошибки совместного доступа
new_wb.close()
# конец выполнения программы