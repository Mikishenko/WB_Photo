# ОБРАБАТЫВАЕТ ИЗ НЕСКОЛЬКИХ ФАЙЛОВ С ОПРЕДЕЛЁННОЙ СТРУКТУРОЙ
# (ИЗ ozon каталога) И ГЕНЕРИРУЕТ ОДИН ФАЙЛ


import openpyxl
import os

# устанавливаем рабочий каталог, где лежат файлы с OZON
# path = os.chdir("C:\\Users\\AlexMiki\\Desktop\\ПРОЕКТЫ\\Переезд OZON to WB\\OZON товары")
# path = os.chdir("F:\\Dropbox\\LeessonS\\WB_Photo\\OZONE")
path = os.chdir("D:\\Dropbox\\LeessonS\\analysis")

# РАБОТА С СОЗДАВАЕМЫМ ФАЙЛОМ
# инициализация НОВОГО Excel-файла и листа для записи
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# создаём заголовки в  первой строке листа
# перечисляем фиксированные имена полей будущей таблицы
table_tytles = ('ID','Наименование','ЦенаМаркетПлейс','Тип','ШК_OZONE','Вес',
               'Высота','Ширина','Глубина','URL_фото','Производитель','Модель',
               'Тип_2')
# записываем поля в файл
for col_num in range (0, 13):
    new_ws.cell(row = 1, column=col_num+1).value = table_tytles[col_num]
#  переменная 'i' - используется в адресации строк (начнём со 2-й строки)
i = 2
# РАБОТА СО СПИСКОМ ИМЕЮЩИХСЯ ФАЙЛОВ

file_count = 0 # счётчик для отображения номера файла, который в работе

# определяем список файлов из каталога и запускаем обработку
for name_file in os.listdir(path):
    file_count +=1
    print (file_count, '\t', name_file) # выводим для информации имя файла
    wb = openpyxl.open(name_file) # открываем книгу, ссылаясь на текущий файл
    wb.active = 4 # устанавливаем активный лист (определён заранее)
    ws = wb.active
    # считываем список значений ячеек с листа, начиная с 4-й строки
    # заголовки пропускаются
    for row in range (4, ws.max_row+1):
        # выполняем проверку на "пустоту" ячейки
        # если не пустая, то считывается значение переменной
        if ws[row][1].value :
            id_oscomp = ws[row][1].value
            name_product = ws[row][2].value
            price = ws[row][3].value
            type_product = ws[row][8].value
            bar_ozone = ws[row][9].value
            weight = ws[row][10].value
            width = ws[row][11].value
            height = ws[row][12].value
            depth = ws[row][13].value
            photo_url = ws[row][14].value
            brand = ws[row][19].value
            model = ws[row][20].value
            type_prod_v2 = ws[row][22].value
            # полученные значения  записываются в столбцы текущей строки (i)
            # у новой рабочей книги 'ws'
            new_ws.cell(row = i, column = 1).value = id_oscomp
            new_ws.cell(row = i, column = 2).value = name_product
            new_ws.cell(row = i, column = 3).value = price
            new_ws.cell(row = i, column = 4).value = type_product
            new_ws.cell(row = i, column = 5).value = bar_ozone
            new_ws.cell(row = i, column = 6).value = weight
            new_ws.cell(row = i, column = 7).value = width
            new_ws.cell(row = i, column = 8).value = height
            new_ws.cell(row = i, column = 9).value = depth
            new_ws.cell(row = i, column = 10).value = photo_url
            new_ws.cell(row = i, column = 11).value = brand
            new_ws.cell(row = i, column = 12).value = model
            new_ws.cell(row = i, column = 13).value = type_prod_v2
        # если ячейка пустая - счётчик строки для создаваемого файла уменьшается
        else:
            break
        # увеличиваем значение строки для записи на следующей строке листа
        i += 1
# по окончании обработки всех файлов сохраняем изменения в созданном файле
new_wb.save("ALL_catalog_from_OZON_in_one_file.xlsx")
# закрываем книгу для исключения ошибки совместного доступа
new_wb.close()
# конец выполнения программы