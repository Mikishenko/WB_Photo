#  приложение генерирует номер серии для указанного в файле количества
# создаёт Excel файл для подготовки прихода товара с сериями в 1С УНФ

# берётся один файл и у тех товаров, где есть индекс "1"
# в колонке с "статусом серии" - формирует уникальный номер серии
# и кладёт в файл с дополненными данными в др. файл


import openpyxl
import os

# устанавливаем рабочий каталог, где лежат файлы с OZON

path = os.chdir("D:\\Dropbox\\LeessonS\\SERIES")

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

table_tytles = ('ID','Name','Count','Price','status', 'series')

for col_num in range (0, 5):
    new_ws.cell(row = 1, column=col_num+1).value = table_tytles[col_num]
#  переменная 'i' - используется в адресации строк (начнём со 2-й строки)
i = 2
# РАБОТА СО СПИСКОМ ИМЕЮЩИХСЯ ФАЙЛОВ
xx = 0
file_count = 0 # счётчик для отображения номера файла, который в работе
series = []
a = 1000 #начальный номер серии
# определяем список файлов из каталога и запускаем обработку
for name_file in os.listdir(path):
    file_count +=1
    print (file_count, '\t', name_file) # выводим для информации имя файла
    wb = openpyxl.open(name_file) # открываем книгу, ссылаясь на текущий файл
    wb.active = 0 # устанавливаем активный лист (определён заранее)
    ws = wb.active
    # считываем список значений ячеек с листа, начиная с 4-й строки
    # заголовки пропускаются

    for row in range (2, ws.max_row+1):
        # выполняем проверку на "пустоту" ячейки
        # если не пустая, то считывается значение переменной
        if ws[row][1].value :
            id_oscomp = ws[row][0].value
            name_product = ws[row][1].value
            count = ws[row][2].value
            price = ws[row][3].value
            status = ws[row][4].value
            # полученные значения  записываются в столбцы текущей строки (i)
            # у новой рабочей книги 'ws'
            new_ws.cell(row = i, column = 1).value = id_oscomp
            new_ws.cell(row = i, column = 2).value = name_product
            new_ws.cell(row = i, column = 3).value = count
            new_ws.cell(row = i, column = 4).value = price
            new_ws.cell(row = i, column = 5).value = status
            if status == 1:
                for xx in range (0, count):
                    series.append("OSC_"+str(a))
                    a = a+1
                    new_ws.cell(row = i, column = 6).value = str(series)
                series = []
        # если ячейка пустая - счётчик строки для создаваемого файла уменьшается
        else:
            break
        # увеличиваем значение строки для записи на следующей строке листа
        i += 1
# по окончании обработки всех файлов сохраняем изменения в созданном файле
new_wb.save("SERIES_with_number.xlsx")
# закрываем книгу для исключения ошибки совместного доступа
new_wb.close()
# конец выполнения программы